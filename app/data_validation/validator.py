import pandas as pd
from openpyxl import load_workbook
import math

# from input_handler.load_parameters import load_configuration_file
from input_handler.load_parameters import load_parameters
from input_handler.data_preprocessor import data_preprocessor

from memory_profiler import profile


class env_setting():
    def __init__(self, context):
        self.run_yymm = context.run_yymm
        self.prtflo_scope = context.prtflo_scope  # self.run_scope = context.run_scope
        self.scenario_version = context.SCENARIO_VERSION
        self.t_zero = context.T_ZERO
        #self.lgd_scope = context.prtflo_scope_fa
        self.scenario_set = context.scenario_set
        self.days_in_year = context.days_in_year
        self.days_in_month = context.days_in_month
        self.total_yr = context.total_yr

        self.masterPath = context.masterPath
        self.parmPath = context.parmPath
        self.dataPathScen = context.dataPathScen
        self.inDataPath = context.inDataPath
        self.resultPath = context.resultPath
        self.dataName = context.dataNameScen  # scenario data

        self.dtype_tbl = context.dtype_tbl
        self.inputDataExtECL = context.inputDataExtECL
        self.instrument_table_name = context.instrument_table_name
        self.exchange_rate_table_name = context.exchange_rate_table_name
        self.repayment_table_name = context.repayment_table_name
        self.sa_fs_table_name = context.sa_fs_table_name
        self.sa_other_debt_table_name = context.sa_other_debt_table_name
        self.mute_eir = context.mute_eir
        self.mute_stage_consistency = context.mute_stage_consistency

class data_health_check_basic(env_setting):
    def __init__(self, context):
        super().__init__(context)

    # *************** General Basic function for loading parameter and raw data ***************
    def load_data(self):
        """
        Load scenario data located at `dataPathScen` with the specified `dataName` and sheet name 'Data'.

        Returns:
        - DataFrame: The raw scenario data loaded from the Excel file.
        """
        raw = pd.read_excel(self.dataPathScen /
                            self.dataName, sheet_name='Data')
        return raw

    def load_raw_data(self, inDataPath, inDataExt, dtype_tbl, param_dict):
        """
        Load instrument table, exchange rate, and repayment csv data located at `inDataPath`, and cast them by defined type.

        Parameters:
        - inDataPath: data stored path
        - inDataExt: the extenstion of data files
        - dtype_tbl: data type list
        - param_dict: loaded parameter files

        Returns:
        - Dict: A dictionary containing processed data from each csv, where the keys are csv names.
        """
        param = param_dict.copy()
        dp = data_preprocessor(context=self)

        exposure_std = dp.standardize_data(inDataPath=inDataPath, rawDataName=self.instrument_table_name,
                                           dtype_tbl=dtype_tbl, param_dict=param, inputDataExt=inDataExt)
        fx_tbl = dp.standardize_data(inDataPath=inDataPath, rawDataName=self.exchange_rate_table_name,
                                     dtype_tbl=dtype_tbl, param_dict=param, inputDataExt=inDataExt)
        # col_tbl = dp.standardize_data(inDataPath=inDataPath, rawDataName=self.collateral_table_name,
        #                               dtype_tbl=dtype_tbl, param_dict=param, inputDataExt=inDataExt)
        cf_tbl = dp.standardize_data(inDataPath=inDataPath, rawDataName=self.repayment_table_name,
                                     dtype_tbl=dtype_tbl, param_dict=param, inputDataExt=inDataExt)

        return {'exposure_std': exposure_std[0],
                'fx_tbl': fx_tbl[0],
                'cf_tbl': cf_tbl[0],
                # 'col_tbl': col_tbl[0], 
                }


class data_health_check(data_health_check_basic):
    def __init__(self, context):
        super().__init__(context)

    def load_basic_info(self):
        """
        Load parameters and data table via call func in data_health_check_basic.

        Returns:
        - tuple: A tuple containing the loaded parameters and raw data table.
        """
        param = load_parameters(parmPath=self.parmPath)
        raw_df_set = self.load_raw_data(inDataPath=self.inDataPath,
                                        inDataExt=self.inputDataExtECL,
                                        dtype_tbl=self.dtype_tbl,
                                        param_dict=param)
        return param, raw_df_set

    def load_data_check_param(self, param):
        """
        Load validation parameters.

        Parameters:
        - param (dict): loaded parameter tables

        Returns:
        - Dict: A dictionary containing processed data from ValidationParam.xlsx, where the keys are sheet names.
        """
        datacheckparm = {}
        _validfile = list(self.parmPath.glob(f'[!~]*ValidationParam.xlsx'))[0]
        _validsheet = load_workbook(_validfile).sheetnames
        for _parm in _validsheet:
            _key = _parm.split(' ')[0]
            datacheckparm[_key] = param[_key]
        return datacheckparm

    # Parameter check: check the existence of all parameter columns in parameter files
    def check_param(self, param_dict, validateparam):
        """
        Check the existence of all parameter tabs in parameter files (AutoLGD, AutoPD, DataValidation, general, model, pwa),
        and check whether the parameter are empty in parameter files

        Parameters:
        - param_dict:loaded parameter tables in dictionary
        - validateparam: DataValidation parameter tables in dictionary

        Returns:
        - list: A list containing the result of the data quality check. If all checks pass, returns ['Pass'].
                If any issues are found, returns ['Fail'] followed by a list of detailed exceptions.
        """

        param = param_dict.copy()
        ParamName = validateparam["ParamName"][validateparam["ParamName"]
                                               ['ParamFile'] != 'DataValidationParam.xlsx'].copy()
        validateparamName = validateparam["ParamName"][validateparam["ParamName"]
                                                       ['ParamFile'] == 'DataValidationParam.xlsx'].copy()
        ParamName = ParamName.reset_index(drop=True)
        validateparamName = validateparamName.reset_index(drop=True)
        lastsheet = ""
        MissingSheet = []
        MissingSheet_File = []
        MissingName = []
        MissingName_Sheet = []
        MissingName_File = []
        MissingData = []
        MissingData_Sheet = []
        MissingData_File = []

        # check parameter tabs and parameter columns in DataValidationParam
        for i in range(len(validateparamName["ParamTab"])):
            sheet = validateparamName["ParamTab"][i].split(' ')[0]
            chk = validateparamName["ParamName"][i]
            if sheet in validateparam:
                if validateparam[sheet].columns.isin([chk]).any() == False:
                    MissingName.append(chk)
                    MissingName_Sheet.append(sheet)
                    MissingName_File.append(validateparamName["ParamFile"][i])
                else:
                    if any(pd.isnull(validateparam[sheet][chk])) == True:
                        MissingData.append(chk)
                        MissingData_Sheet.append(sheet)
                        MissingData_File.append(
                            validateparamName["ParamFile"][i])
            else:
                if lastsheet != sheet:
                    MissingSheet.append(sheet)
                    MissingSheet_File.append(validateparamName["ParamFile"][i])
            lastsheet = sheet

        lastsheet = ""
        # check parameter tabs and parameter columns in *Param files
        for i in range(len(ParamName["ParamTab"])):
            # for "instrument_table", keep the "A_B" structure
            sheet = ParamName["ParamTab"][i].split(' ')[0]
            chk = ParamName["ParamName"][i]
            if sheet in param:
                if param[sheet].columns.isin([chk]).any() == False:
                    MissingName.append(chk)
                    MissingName_Sheet.append(sheet)
                    MissingName_File.append(ParamName["ParamFile"][i])
                else:
                    # 20240130: Bug fixed as any()
                    if any(pd.isnull(param[sheet][chk])) == True:
                        MissingData.append(chk)
                        MissingData_Sheet.append(sheet)
                        MissingData_File.append(ParamName["ParamFile"][i])
            else:
                if lastsheet != sheet:
                    MissingSheet.append(sheet)
                    MissingSheet_File.append(ParamName["ParamFile"][i])
            lastsheet = sheet

        if len(MissingName) == 0 and len(MissingSheet) == 0 and len(MissingData) == 0:
            return ['Pass']
        else:
            paramException = []
            if len(MissingSheet) > 0:
                for i in range(len(MissingSheet)):
                    paramException.append(
                        'Tab "' + str(MissingSheet[i]) + '" is missing in "' + str(MissingSheet_File[i]))
            elif len(MissingName) > 0:
                for i in range(len(MissingName)):
                    paramException.append('Parameter "' + str(MissingName[i]) + '" is missing in tab "' + str(
                        MissingName_Sheet[i]) + '" of ' + str(MissingName_File[i]))
            else:
                for i in range(len(MissingData)):
                    paramException.append('Parameter "' + str(MissingData[i]) + '" is empty in tab "' + str(
                        MissingData_Sheet[i]) + '" of ' + str(MissingData_File[i]))
            return ['Fail'] + paramException

# Duplicate key check
    def check_dup_key(self, input_df, id_col):
        """
        Check whether input data contains duplicated contract id.

        Parameters:
        - input_df: instrument table
        - id_col(str): key name

        Returns:
        - list: A list containing the result of the dup key check. If all checks pass, returns ['Pass'].
                If any issues are found, returns ['Fail'] followed by a list of detailed exceptions.
        """
        dup_rec = []
        # dup_rec = input_df[input_df[id_col].duplicated()][id_col]
        dup_rec_on = input_df[input_df['ON_OFF_BAL_IND'] ==
                              'ON'][input_df[input_df['ON_OFF_BAL_IND'] == 'ON'][id_col].duplicated()][id_col]
        dup_rec_off = input_df[input_df['ON_OFF_BAL_IND'] ==
                               'OFF'][input_df[input_df['ON_OFF_BAL_IND'] == 'OFF'][id_col].duplicated()][id_col]
        # dup_rec_all = input_df[[id_col].duplicated()][id_col]
        dup_rec.append('on:')
        for item in dup_rec_on:
            dup_rec.append(item)
        dup_rec.append('off:')
        for item in dup_rec_off:
            dup_rec.append(item)
        if (len(dup_rec_off)+len(dup_rec_on) == 0):
            return ['Pass']
        else:
            return ['Fail', dup_rec]

# Parameter compeleteness check
    def check_param_completeness(self, param_dict, fullListParam):
        """
        Check whether whether input data contains complete set of parameter required for ECL engine run.

        Parameters:
        - param_dict:loaded parameter tables in dictionary
        - fullListParam: the dataframe of 'ValidParam' in validate parm listed all parameter table names

        Returns:
        - list: A list containing the result of the parameter table list completeness check. If all checks pass, returns ['Pass'].
                If any issues are found, returns ['Fail'] followed by a list of detailed exceptions.
        """
        param_set = set(list(param_dict.keys()))
        validparam_set = set(fullListParam['ParamName'])
        if param_set == validparam_set:
            return ['Pass']
        else:
            paramException = {'missing': list(validparam_set - param_set),
                              'extra': list(param_set - validparam_set)}
            return ['Fail', paramException]

# Check data consistency
    def check_data_consistency(self, param_dict, raw_df_dict):
        """
        Check whether exchange rate, and prepayment table contain all necessary information required for ECL engine run.

        Parameters:
        - param_dict:loaded parameter tables in dictionary
        - raw_df_dict: the dict of data

        Returns:
        - list: A list containing the result of the data consistency check. If all checks pass, returns ['Pass'].
                If any issues are found, returns ['Fail'] followed by a list of detailed exceptions.
        """
        param = param_dict.copy()
        raw_df = raw_df_dict.copy()

        logic_list = []
        # all ocy are includeded in exchange_rate_table.csv
        CCY_instrument = set(raw_df['exposure_std'].CCY_CD)
        CCY_exchange = set(raw_df['fx_tbl'].CCY_CD_OCY)
        fx_pass_ind = CCY_instrument.issubset(CCY_exchange)

        logic_list.append(fx_pass_ind)

        # all interest rate index code are included in interestRateInput.csv
        # int_pass_ind = set(set(raw_df['exposure_std'].EFF_INT_RT)-set(['0','nan'])).issubset(raw_df['int_rt_tbl'].INT_RT_IDX)
        # logic_list.append(int_pass_ind)

        # all cashflow CONTRACT_ID on balance are included in repayment_table.csv
        # ID_instrument = set(raw_df['exposure_std'].query('ON_OFF_BAL_IND=="ON"').CONTRACT_ID)

        # only loans will do collective assessemnt and specific assessment
        dp = data_preprocessor(context=self)
        instr_df = raw_df['exposure_std']
        instr_df_in_scope = instr_df.query(
            "DATA_SOURCE_CD.isin(@self.prtflo_scope)")  # run_scope
        # Map credit notch number
        instr_df_1 = dp._map_credit_notch(df=instr_df_in_scope, param=param)
        # Perform Stage allocation
        instr_df_2 = dp.allocate_stage(df=instr_df_1,
                                         param=param)
        # instr_df_1 = dp._get_facility_stage(df=instr_df_filter)
        # instr_df_2 = dp._get_customer_stage(df=instr_df_1)
        # instr_df_3 = dp._get_overlaid_stage(df=instr_df_2, param=param)
        # instr_df_4 = dp._get_final_stage(df=instr_df_3)
        instr_df_3 = dp.allocate_ecl_approach(df=instr_df_2, param=param)

        report_date = instr_df_3.REPORT_DATE[1]
        ID_instrument = set(instr_df_3.query(
            "DATA_SOURCE_CD == 'LOAN' and ECL_APPROACH != 'PROXY' and MAT_DATE > @report_date and ON_OFF_BAL_IND == 'ON'").CONTRACT_ID)

        ID_repayment = set(raw_df['cf_tbl'].CONTRACT_ID)
        cf_pass_ind = ID_instrument.issubset(ID_repayment)
        logic_list.append(cf_pass_ind)

        if all(logic_list):
            return ['Pass']
        else:
            paramException = []
            if fx_pass_ind is False:
                missingCCY = CCY_instrument - CCY_exchange
                paramException.append(
                    {"Missing OCY in exchange_rate_table.csv": list(missingCCY)})
            # if int_pass_ind is False:
            #     paramException.append("Missing interest rate index code in InterestRateInput.csv")
            if cf_pass_ind is False:
                missingID = ID_instrument - ID_repayment
                paramException.append(
                    {"Missing contract id in repayment_table.csv": list(missingID)})
            return ['Fail', paramException]

#############
    # helper func
    def _is_missing(self, x):
        """
        Check if a value is considered as missing.

        Args:
        - x: The value to be checked for missingness.

        Returns:
        - bool: True if the value is missing, False otherwise.
        """
        if (((isinstance(x, int) or isinstance(x, float)) and math.isnan(float(x))) or
                (isinstance(x, str) and (x in ('nan', '', 'NAN')))):
            return True
        else:
            return False

    def _is_all_numeric(self, x):
        """
        Check if all values in the input data are numeric (integers or floats).

        Args:
        - x (iterable): List or iterable containing values to be checked.    
        Returns:
        - bool: True if all values are numeric, False otherwise.
        """
        return all(isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '', 1).isdigit()) for value in x)

    # UNIT TEST
    # 20240408: No rounding to 2 units for missing rate
    def _unit_missing(self, input_df):
        """
        Compute the missing count and missing rate of an array or a list of values.

        Parameters: 
        - input_df: an array or a list of values

        Returns:
        - tuple: A tuple containing the sum number of missing count and the missing rate.
        """
        miss_cnt = input_df.apply(lambda x: self._is_missing(x)).sum()
        return (miss_cnt, miss_cnt / input_df.shape[0])

    def _unit_numerical_date_var(self, input_df):
        """
        Compute the min, max, median of an array or a list of values.

        Parameters: 
        - input_df: an array or a list of values

        Returns:
        - tuple: A tuple containing the min, max, median
        """
        minOut = 0.0 if input_df.min() is None else input_df.min()
        maxOut = 0.0 if input_df.max() is None else input_df.max()
        medianOut = 0.0 if input_df.median() is None else input_df.median()
        return (minOut, maxOut, medianOut)

    def _unit_categorical_var(self, input_df):
        """
        Return non-missing unique value of an array of a list.

        Parameters: 
        - input_df: an array or a list of values

        Returns:
        - list: A list containing unique value without missing elements
        """
        # only return if not missing
        uniqlist = input_df.unique().tolist()
        uniqlist_rm = [x for x in uniqlist if self._is_missing(x) == False]
        return uniqlist_rm
# MISSING CHECK

    def check_missing(self, ScopeofVar, raw_df_dict):
        """
        Return the missing statistics of the instrument table:
            1. report missing rate of each column by DATA_SOURCE_CD only
            2. go through missing test if needed
            If there is No missing / non check item missing: return GREEN
            otherwise, any missing item: return RED
            #Immaterial missing but worth to mention: AMBER

        Parameters: 
        - ScopeofVar: the 'ScopeofVar' parameter in validation parameter file
        - input_df: inputted data list

        Returns:
        - tuple: A list containing the result of the data missing check. If all checks pass, returns ['Pass'].
                If any issues are found, returns ['Fail'] followed by a list of detailed exceptions.
        """
        raw_df = raw_df_dict['exposure_std'].copy()
        report_col_list = ['DATA_SOURCE_CD', 'INSTR_TYPE',
                           'VARIABLE', 'MISS_CNT', 'MISS_RATE', 'CHECK_FLG']
        MissingReport = pd.DataFrame(columns=report_col_list)
        chk_flg_dist = {0: 'GREEN', 1: 'AMBER', 2: 'RED'}
        # Check if DATA_SOURCE_CD is missing
        miss_DATA_SOURCE_CD, miss_rate_DATA_SOURCE_CD = self._unit_missing(
            raw_df['DATA_SOURCE_CD'])
        if miss_DATA_SOURCE_CD > 0:
            res_list = ['N/A', 'N/A', 'DATA_SOURCE_CD',
                        miss_DATA_SOURCE_CD, miss_rate_DATA_SOURCE_CD, 1]
            tmp = pd.DataFrame([dict(zip(report_col_list, res_list))])
            MissingReport = pd.concat(
                [MissingReport, tmp], axis=0, ignore_index=True)
        # if miss_INSTR_TYPE > 0:
        #     res_list = ['N/A','N/A','INSTR_TYPE',miss_INSTR_TYPE,miss_rate_INSTR_TYPE,1]
        #     tmp = pd.DataFrame([dict(zip(report_col_list,res_list))])
        #     MissingReport = pd.concat([MissingReport, tmp], axis=0, ignore_index=True)
        #     #TO CHECK: if want to drop N/A DATA_SOURCE_CD, INSTR_TYPE rows

        IndexPairList = raw_df.groupby(
            ['DATA_SOURCE_CD', 'INSTR_TYPE']).count().index.to_frame(index=False)
        chk_flg = -1
        for index, row in IndexPairList.iterrows():
            tgt_DATA_SOURCE_CD = row['DATA_SOURCE_CD']
            tgt_INSTR_TYPE = row['INSTR_TYPE']
            raw_df_filter = raw_df.query(
                'DATA_SOURCE_CD == @tgt_DATA_SOURCE_CD and INSTR_TYPE == @tgt_INSTR_TYPE').copy()

            # generate missing rate for each column
            for col in raw_df_filter.columns.tolist():
                miss_cnt, miss_rate = self._unit_missing(raw_df_filter[col])
                if miss_rate == 0.0:
                    chk_flg = 0
                else:
                    scopedf = ScopeofVar.query(
                        'DATA_SOURCE_CD == @tgt_DATA_SOURCE_CD and varname == @col and Function == "Check_missing"').copy()
                    # Criteria: missing is acceptable if under this criteria
                    if scopedf.shape[0] == 1:
                        scopedf = scopedf.reset_index()
                        if pd.isnull(scopedf.loc[0, 'Criteria']):
                            chk_flg = 2
                        else:
                            # all convert to upper case for comparison
                            raw_df_filter_up = raw_df_filter.apply(
                                lambda x: x.astype(str).str.upper())
                            criteria_value = scopedf.loc[scopedf.index[0], 'Criteria']
                            if criteria_value == 'nan':
                                raw_df_exclude = pd.DataFrame(
                                    columns=raw_df_filter_up.columns)
                            else:
                                raw_df_exclude = raw_df_filter_up.query(
                                    criteria_value).copy()

                            if raw_df_exclude.shape[0] == 0:
                                chk_flg = 2
                            else:
                                miss_cnt_exclude, miss_rate_exclude = self._unit_missing(
                                    raw_df_exclude[col])
                                if miss_cnt_exclude == miss_cnt:
                                    chk_flg = 0
                                else:
                                    chk_flg = 2
                    else:
                        # if not in check list and have missing then report GREEN
                        chk_flg = 0
                res_list = [tgt_DATA_SOURCE_CD, tgt_INSTR_TYPE,
                            col, miss_cnt, miss_rate, chk_flg]
                # res_list = [tgt_DATA_SOURCE_CD, col, miss_cnt, miss_rate, chk_flg]
                tmp = pd.DataFrame([dict(zip(report_col_list, res_list))])
                MissingReport = pd.concat(
                    [MissingReport, tmp], axis=0, ignore_index=True)
        missing_check_final_ind = chk_flg_dist[max(MissingReport['CHECK_FLG'])]
        MissingReport['CHECK_FLG'] = MissingReport['CHECK_FLG'].map(
            chk_flg_dist)
        return (missing_check_final_ind, MissingReport)

# VALID VALUE STATISTICS & CHECK
    # CATEGORICAL: if
    def check_categorical_stats(self, ScopeofVar, ValidCat, raw_df_dict):
        """
        Check if the unique value in all category data falls in valid category list.
        All unique value in valid list then pass else fail, record in exception.

        Parameters: 
        - ScopeofVar: the 'ScopeofVar' parameter in validation parameter file.
        - ValidCat: valid category list.
        - raw_df_dict: inputted data list

        Returns:
        - tuple: A list containing the result of the data category check. If all checks pass, returns ['Pass'].
                If any issues are found, returns ['Fail'] followed by a list of detailed exceptions.
        """
        raw_df = raw_df_dict['exposure_std'].copy()
        report_col_list = ['DATA_SOURCE_CD',
                           'INSTR_TYPE', 'VARIABLE', 'VALUES', 'CHK_FLG']
        StatsReport = pd.DataFrame(columns=report_col_list)
        exception_col_list = ['DATA_SOURCE_CD', 'INSTR_TYPE',
                              'VARIABLE', 'INPUT_NOT_VALID', 'CHK_FLG']
        ExceptionReport = pd.DataFrame(columns=exception_col_list)
        chk_flg_dist = {0: 'Pass', 1: 'Fail'}
        IndexPairList = raw_df.groupby(
            ['DATA_SOURCE_CD', 'INSTR_TYPE']).count().index.to_frame(index=False)
        chk_flg = -1
        for index, row in IndexPairList.iterrows():
            tgt_DATA_SOURCE_CD = row['DATA_SOURCE_CD']
            tgt_INSTR_TYPE = row['INSTR_TYPE']
            raw_df_filter = raw_df.query(
                'DATA_SOURCE_CD == @tgt_DATA_SOURCE_CD and INSTR_TYPE == @tgt_INSTR_TYPE').copy()
            # for col in raw_df_filter.select_dtypes(include='object').columns.tolist():
            for col in ValidCat['varname'].unique().tolist():
                # convert both to upper() to compare
                distinctValList = [
                    str(x).upper() for x in self._unit_categorical_var(raw_df_filter[col])]
                # scopedf = ScopeofVar.query('DATA_SOURCE_CD == @tgt_DATA_SOURCE_CD and INSTR_TYPE == @tgt_INSTR_TYPE and varname == @col and Function == "Check_distinct"')
                scopedf = ScopeofVar.query(
                    'DATA_SOURCE_CD == @tgt_DATA_SOURCE_CD and varname == @col and Function == "Check_distinct"').copy()
                if scopedf.shape[0] == 0:
                    chk_flg = 0

                # TO CHECK: automatically fail more than 1 range check?
                elif scopedf.shape[0] > 1:
                    chk_flg = 1
                    disc_list = [tgt_DATA_SOURCE_CD, tgt_INSTR_TYPE, [
                        'More than 1 criteria found'], chk_flg]
                    tmpExecp = pd.DataFrame(
                        [dict(zip(exception_col_list, disc_list))])
                    ExceptionReport = pd.concat(
                        [ExceptionReport, tmpExecp], axis=0, ignore_index=True)
                else:
                    # get valid categorical values,convert both to upper() to compare
                    validList = [str(x).upper() for x in ValidCat.query(
                        'varname == @col')['category'].copy().tolist()]
                    if set(distinctValList).issubset(set(validList)):
                        chk_flg = 0
                    else:
                        chk_flg = 1
                        discrepancy = list(set(distinctValList)-set(validList))
                        disc_list = [tgt_DATA_SOURCE_CD,
                                     tgt_INSTR_TYPE, col, discrepancy, chk_flg]
                        tmpExecp = pd.DataFrame(
                            [dict(zip(exception_col_list, disc_list))])
                        ExceptionReport = pd.concat(
                            [ExceptionReport, tmpExecp], axis=0, ignore_index=True)

                report_res = [tgt_DATA_SOURCE_CD, tgt_INSTR_TYPE,
                              col, distinctValList, chk_flg]
                tmpStats = pd.DataFrame(
                    [dict(zip(report_col_list, report_res))])
                StatsReport = pd.concat(
                    [StatsReport, tmpStats], axis=0, ignore_index=True)

        stat_check_final_ind = chk_flg_dist[max(StatsReport['CHK_FLG'])]
        StatsReport['CHK_FLG'] = StatsReport['CHK_FLG'].map(chk_flg_dist)
        return (stat_check_final_ind, StatsReport, ExceptionReport)

    # NUMERICAL: if min, max in range then pass else fail, record in exceoption
    def check_numerical_stats(self, ScopeofVar, raw_df_dict):
        """
        Check if the value of varaibles can pass the defined range checking.
        If min or max fall in range, then pass, else fail, record in exception.

        Parameters: 
        - ScopeofVar: the 'ScopeofVar' parameter in validation parameter file.
        - raw_df_dict: inputted data list

        Returns:
        - tuple: A list containing the result of the data range check. If all checks pass, returns ['Pass'].
                If any issues are found, returns ['Fail'] followed by a list of detailed exceptions.
        """
        raw_df = raw_df_dict['exposure_std'].copy()
        report_col_list = ['DATA_SOURCE_CD', 'INSTR_TYPE',
                           'VARIABLE', 'MIN', 'MAX', 'MEDIAN', 'CHK_FLG']
        StatsReport = pd.DataFrame(columns=report_col_list)
        exception_col_list = ['DATA_SOURCE_CD', 'INSTR_TYPE',
                              'VARIABLE', 'FAIL_CRITERIA', 'CHK_FLG']
        ExceptionReport = pd.DataFrame(columns=exception_col_list)
        chk_flg_dist = {0: 'Pass', 1: 'Fail'}
        IndexPairList = raw_df.groupby(
            ['DATA_SOURCE_CD', 'INSTR_TYPE']).count().index.to_frame(index=False)
        chk_flg = -1
        for index, row in IndexPairList.iterrows():
            tgt_DATA_SOURCE_CD = row['DATA_SOURCE_CD']
            tgt_INSTR_TYPE = row['INSTR_TYPE']
            raw_df_filter = raw_df.query(
                'DATA_SOURCE_CD == @tgt_DATA_SOURCE_CD and INSTR_TYPE == @tgt_INSTR_TYPE').copy()
            for col in raw_df_filter.select_dtypes(include=['number', 'datetime']).columns.tolist():
                minVal, maxVal, medianVal = self._unit_numerical_date_var(
                    raw_df_filter[col])
                # scopedf = ScopeofVar.query('DATA_SOURCE_CD == @tgt_DATA_SOURCE_CD and INSTR_TYPE == @tgt_INSTR_TYPE and varname == @col and Function == "check_range"')
                scopedf = ScopeofVar.query(
                    'DATA_SOURCE_CD == @tgt_DATA_SOURCE_CD and varname == @col and Function == "check_range"').copy()
                if scopedf.shape[0] == 0:
                    chk_flg = 0

                # TO CHECK: automatically fail more than 1 range check?
                elif scopedf.shape[0] > 1:
                    chk_flg = 1
                    disc_list = [tgt_DATA_SOURCE_CD, tgt_INSTR_TYPE, [
                        'More than 1 criteria found'], chk_flg]
                    tmpExecp = pd.DataFrame(
                        [dict(zip(exception_col_list, disc_list))])
                    ExceptionReport = pd.concat(
                        [ExceptionReport, tmpExecp], axis=0, ignore_index=True)
                else:
                    # decompose criteria, excel need to break rules with semicolon
                    chk_criteria = [x.strip()
                                    for x in scopedf['Critermiia'].split(';')]
                    res = [eval(i) for i in chk_criteria]
                    if all(res):
                        chk_flg = 0
                    else:
                        chk_flg = 1
                        chk_dict = dict(zip(chk_criteria, res))
                        for key, val in chk_dict.item():
                            if val == 0:
                                disc_list = [tgt_DATA_SOURCE_CD,
                                             tgt_INSTR_TYPE, col, key, val]
                                tmpExecp = pd.DataFrame(
                                    [dict(zip(exception_col_list, disc_list))])
                                ExceptionReport = pd.concat(
                                    [ExceptionReport, tmpExecp], axis=0, ignore_index=True)

                report_res = [tgt_DATA_SOURCE_CD, tgt_INSTR_TYPE,
                              col, minVal, maxVal, medianVal, chk_flg]
                tmpStats = pd.DataFrame(
                    [dict(zip(report_col_list, report_res))])
                StatsReport = pd.concat(
                    [StatsReport, tmpStats], axis=0, ignore_index=True)

        stat_check_final_ind = chk_flg_dist[max(StatsReport['CHK_FLG'])]
        StatsReport['CHK_FLG'] = StatsReport['CHK_FLG'].map(chk_flg_dist)
        return (stat_check_final_ind, StatsReport, ExceptionReport)

    # FIXME: The checking need to talk to the scenario module
# Scenario checking:
    def check_scenario(self, dv_param_MEF, raw_scen):
        """
        Check if the scenario data:
            1. check whether MEF are in correct order in terms of name and scenario
            2. check MEF data at t0 quarter are numeric
            3. check MEF data a year later after t0 quarter are numeric

        Parameters: 
        - dv_param_MEF: the 'MEFname' parameter in validation parameter file.
        - raw_scen: scenario data

        Returns:
        - tuple: A list containing the result of the scenario data check. If all checks pass, returns ['Pass'].
                If any issues are found, returns ['Fail'] followed by a list of detailed exceptions.
        """
        param_mef = dv_param_MEF.copy()
        logic_list = []

        # 1. check MEF names and scenarios in correct order, return wrong order MEF
        MD_col_name = param_mef['MEF_code'][0:].to_list()
        raw_code = raw_scen.columns[1:].tolist()
        scenario_check_ind = raw_code == MD_col_name
        logic_list.append(scenario_check_ind)

        #2. check whether the value in row of t0 quater (ex, 2024Q1) is not float type number
        quarter_obj = pd.Period(self.t_zero, freq='Q').strftime('%YQ%q')
        #print(quarter_obj)
        scenario_row = raw_scen[raw_scen.iloc[:, 0].str.startswith(quarter_obj)]
        scenario_version_check_ind = self._is_all_numeric(scenario_row.values[0,1:])
        #scenario_version_check_ind = not (scenario_row == 'ND').any().any()
        logic_list.append(scenario_version_check_ind)

        #3. check whether the value of 1yr later is not "ND"
        sc_yr = quarter_obj[:4]
        sc_quarter = quarter_obj[4:]
        sc_yr_1 = str(int(sc_yr) + 1)
        SCENARIO_VERSION_1yr = sc_yr_1 + sc_quarter
        scenario_row_1yr = raw_scen[raw_scen.iloc[:,
                                                  0].str.startswith(SCENARIO_VERSION_1yr)]
        scenario_version_1yr_check_ind = self._is_all_numeric(
            scenario_row_1yr.values[0, 1:])
        # scenario_version_1yr_check_ind = not (scenario_row_1yr == 'ND').any().any()
        logic_list.append(scenario_version_1yr_check_ind)
        # Compile all the logic check result
        if all(logic_list):
            return ['Pass']
        else:
            paramException = []
            if scenario_check_ind is False:
                paramException.append(
                    "The MEF order of raw scenario file does not match with ECL model default column order ")
            if scenario_version_check_ind is False:
                paramException.append(f"The MEF values exist non-number at row {quarter_obj}")
            if scenario_version_1yr_check_ind is False:
                paramException.append(
                    f"The MEF values exist non-number at row {SCENARIO_VERSION_1yr}")
            return ['Fail'] + paramException


class data_health_check_report(data_health_check):
    def __init__(self, context):
        super().__init__(context)

    def data_health_check_report_out(self):
        """
        The data health check report generating function.

        Returns:
        - tuple: A list containing the result of pre run validation check.
        """
        outdf_dict = {}
        # run preparation steps
        dv_scen = self.load_data()
        dv_param, dv_raw_df_set = self.load_basic_info()
        dv_validateparm = self.load_data_check_param(param=dv_param)
        # run all validation
        try:
            stepinfo = "Scenario data checking"
            print(stepinfo)
            scenario_check_res = self.check_scenario(
                dv_param['MEFname'], dv_scen)
        except Exception as e:
            print(f"An error occured when running {stepinfo}: {e}")

        stepinfo = "Parameter existence checking"
        print(stepinfo)
        param_check_res = self.check_param(dv_param, dv_validateparm)

        if param_check_res[0] == "Pass":
            try:
                stepinfo = "Duplicate key checking"
                print(stepinfo)
                # dup_key_res = self.check_dup_key(dv_raw_df_set['exposure_std'][dv_raw_df_set['exposure_std']['ON_OFF_BAL_IND']=='OFF'], 'CONTRACT_ID')
                dup_key_res = self.check_dup_key(
                    dv_raw_df_set['exposure_std'], 'CONTRACT_ID')
            except Exception as e:
                print(f"An error occured when running {stepinfo}: {e}")
                # sys.exit("Stopping the program due to an error")
            try:
                stepinfo = "Data consistency checking"
                print(stepinfo)
                data_consistency_res = self.check_data_consistency(
                    dv_param, dv_raw_df_set)  # , InconsistencyReport
            except Exception as e:
                print(f"An error occured when running {stepinfo}: {e}")
            try:
                stepinfo = "Paramemeter completeness checking"
                print(stepinfo)
                param_completeness_res = self.check_param_completeness(
                    dv_param, dv_validateparm['ValidParam'])
            except Exception as e:
                print(f"An error occured when running {stepinfo}: {e}")
            try:
                stepinfo = "Data missing checking"
                print(stepinfo)
                missing_check_final_ind, MissingReport = self.check_missing(
                    dv_validateparm['ScopeOfVar'], dv_raw_df_set)
            except Exception as e:
                print(f"An error occured when running {stepinfo}: {e}")
            try:
                stepinfo = "Data category checking"
                print(stepinfo)
                stat_check_final_ind_cat, StatsReport_cat, ExceptionReport_cat = self.check_categorical_stats(dv_validateparm['ScopeOfVar'],
                                                                                                              dv_validateparm['ValidCat'], dv_raw_df_set)
            except Exception as e:
                print(f"An error occured when running {stepinfo}: {e}")
            try:
                stepinfo = "Data numeric value checking"
                print(stepinfo)
                stat_check_final_ind_num, StatsReport_num, ExceptionReport_num = self.check_numerical_stats(
                    dv_validateparm['ScopeOfVar'], dv_raw_df_set)
            except Exception as e:
                print(f"An error occured when running {stepinfo}: {e}")

        # concat to different sheets
        # overall result
        validResCols = ['Data Validation Test', 'Result Flag']
        if param_check_res[0] == "Pass":
            validresall = [['Duplicate Key Check', dup_key_res[0]],
                           ['Parameter Completeness Check',
                               param_completeness_res[0]],
                           ['Data Consistency Check', data_consistency_res[0]],
                           ['Missing Rate Check', missing_check_final_ind],
                           ['Data Validity Check (Categorical Variables)',
                            stat_check_final_ind_cat],
                           ['Data Validity Check (Numerical Variables)',
                            stat_check_final_ind_num],
                           ['Scenario Check', scenario_check_res[0]],
                           ['Parameter Check', param_check_res[0]]]
        else:
            validresall = [['Duplicate Key Check', "Parameter is missing, please refer to failReport"],
                           ['Parameter Completeness Check',
                               "Parameter is missing, please refer to failReport"],
                           ['Data Consistency Check',
                            "Parameter is missing, please refer to failReport"],
                           ['Missing Rate Check',
                               "Parameter is missing, please refer to failReport"],
                           ['Data Validity Check (Categorical Variables)',
                            "Parameter is missing, please refer to failReport"],
                           ['Data Validity Check (Numerical Variables)',
                            "Parameter is missing, please refer to failReport"],
                           ['Scenario Check', scenario_check_res[0]],
                           ['Parameter Check', param_check_res[0]]]
        validReport = pd.DataFrame(data=validresall, columns=validResCols)
        # concat with test description
        validReport = validReport.merge(
            dv_validateparm['ValidTestDesc'].iloc[:, 1:], how='left', on='Data Validation Test')
        outdf_dict['validReport'] = validReport

        # failed test & exception
        FailResCols = ['Data Validation Test', 'Result Flag', 'Fail Reason']
        exceptionresall = []
        if len(scenario_check_res) > 1:
            for i in range(1, len(scenario_check_res)):
                exceptionresall.append(
                    ['Scenario Check', scenario_check_res[0], scenario_check_res[i]])
        if len(param_check_res) > 1:
            for i in range(1, len(param_check_res)):
                exceptionresall.append(
                    ['Parameter Check', param_check_res[0], param_check_res[i]])
        else:
            if len(param_completeness_res) > 1:
                exceptionresall.append(
                    ['Parameter Completeness Check', param_completeness_res[0], param_completeness_res[1]])
            if len(dup_key_res) > 1:
                exceptionresall.append(
                    ['Duplicate Key Check', dup_key_res[0], dup_key_res[1]])
            if len(data_consistency_res) > 1:
                exceptionresall.append(
                    ['Data Consistency Check', data_consistency_res[0], data_consistency_res[1]])

        failReport = pd.DataFrame(data=exceptionresall, columns=FailResCols)
        outdf_dict['failReport'] = failReport

        if param_check_res[0] == "Pass":
            # stats
            outdf_dict['MissingReport'] = MissingReport
            outdf_dict['StatsReport_cat'] = StatsReport_cat
            outdf_dict['StatsReport_num'] = StatsReport_num

            # stats exception
            outdf_dict['ExceptionReport_cat'] = ExceptionReport_cat
            outdf_dict['ExceptionReport_num'] = ExceptionReport_num

        # Add sum of record counts and outstanding balance for input file reference
        agg_df = dv_raw_df_set['exposure_std'].groupby('DATA_SOURCE_CD').aggregate({'CONTRACT_ID': 'count',
                                                                                    'PRIN_BAL_LCY': 'sum',
                                                                                    'ACRU_INT_LCY': 'sum',
                                                                                    'PENALTY_LCY': 'sum',
                                                                                    'OTHER_FEE_AND_CHARGES_LCY': 'sum'
                                                                                    }).reset_index()
        agg_df.columns = ['DATA_SOURCE_CD', 'CONTRACT CNT', 'PRIN_BAL_LCY SUM',
                          'ACRU_INT_LCY SUM', 'PENALTY_LCY SUM', 'OTHER_FEE_AND_CHARGES_LCY SUM']
        outdf_dict['DistributionChk'] = agg_df
        return outdf_dict

    # Validation result
    # @profile

    def run(self):
        return self.data_health_check_report_out()
