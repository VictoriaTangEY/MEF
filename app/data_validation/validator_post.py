import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import numpy as np

# from input_handler.env_setting import run_setting
# from input_handler.load_parameters import load_configuration_file
from input_handler.load_parameters import load_parameters
# from input_handler.data_preprocessor import data_preprocessor
# from data_validation.validator import data_health_check

from memory_profiler import profile


class env_setting():
    def __init__(self, context):
        self.run_yymm = context.run_yymm
        self.prtflo_scope = context.prtflo_scope  # self.run_scope = context.run_scope
        self.scenario_version = context.SCENARIO_VERSION
        # self.lgd_scope = context.prtflo_scope_fa
        self.scenario_set = context.scenario_set
        self.days_in_year = context.days_in_year
        self.days_in_month = context.days_in_month
        self.total_yr = context.total_yr

        self.masterPath = context.masterPath
        self.parmPath = context.parmPath
        self.dataPathScen = context.dataPathScen
        self.inDataPath = context.inDataPath
        self.resultPath = context.resultPath
        self.dataName = context.dataNameScen  # scenario data

        self.dtype_tbl = context.dtype_tbl
        self.inputDataExtECL = context.inputDataExtECL
        self.instrument_table_name = context.instrument_table_name
        self.exchange_rate_table_name = context.exchange_rate_table_name
        self.repayment_table_name = context.repayment_table_name
        self.sa_fs_table_name = context.sa_fs_table_name
        self.sa_other_debt_table_name = context.sa_other_debt_table_name


class post_run_validation_basic(env_setting):
    def __init__(self, context):
        super().__init__(context)

    def load_ecl_result(self):
        """
        Load ecl result data located at `resultPath`.

        Returns:
        - DataFrame: The ecl result data loaded from the csv file.
        """
        df_ecl_ = (pd.read_csv(self.resultPath /
                   'ECL_calculation_result_files_deal_all.csv',dtype={'CONTRACT_ID': str, 'CUST_ID':str}))

        cols_str = ['CONTRACT_ID', 'CUST_ID','DATA_SOURCE_CD', 'ON_OFF_BAL_IND']

        for col in cols_str:
            df_ecl_[col] = df_ecl_[col].astype(str)
        return df_ecl_

    def load_instr_data(self):
        """
        Load essensial instrument table located at `inDataPath`.

        Returns:
        - DataFrame: The instrument table with essensial cols loaded from the csv file.
        """

        keep_cols = ['CONTRACT_ID', 'DATA_SOURCE_CD', 'ON_OFF_BAL_IND', 'DRAWN_BAL_OCY', 'UNDRAWN_BAL_OCY',
                     'PRIN_BAL_OCY', 'ACRU_INT_OCY', 'PENALTY_OCY',
                     'OTHER_FEE_AND_CHARGES_OCY']
        df_instr_ = pd.read_csv(f'{self.inDataPath}/{self.instrument_table_name}.{self.inputDataExtECL}',
                                dtype={'CONTRACT_ID': str, 'CUST_ID':str},
                                usecols=keep_cols)
        cols_str = ['CONTRACT_ID', 'DATA_SOURCE_CD', 'ON_OFF_BAL_IND']
        for col in cols_str:
            df_instr_[col] = df_instr_[col].astype(str)

        df_instr_2_ = df_instr_[
            df_instr_['DATA_SOURCE_CD'].str.upper().isin(self.prtflo_scope)].copy()

        return df_instr_2_

    def _load_valid_param(self, param):
        validparm = {}
        _validfile = list(self.parmPath.glob(f'[!~]*ValidationParam.xlsx'))[0]
        _sheetList = load_workbook(_validfile).sheetnames
        for _parm in _sheetList:
            _key = _parm.split('_')[-1]
            validparm[_key] = pd.read_excel(_validfile, sheet_name=_parm)
        return validparm

    def _load_model_param(self, param):
        modelparm = {}
        _modelfile = list(self.parmPath.glob(f'modelParam.xlsx'))[0]
        _sheetList = load_workbook(_modelfile).sheetnames
        for _parm in _sheetList:
            _key = _parm.split('_')[-1]
            modelparm[_key] = pd.read_excel(_modelfile, sheet_name=_parm)
        return modelparm

    def _load_overlay_parm(self, param):
        ovlparm = {}
        _overlayfile = list(self.parmPath.glob(f'overlayTemplate.xlsx'))[0]
        _sheetList = load_workbook(_overlayfile).sheetnames
        for _parm in _sheetList:
            _key = _parm.split('_')[-1]
            ovlparm[_key] = pd.read_excel(_overlayfile, sheet_name=_parm)
        return ovlparm

    def load_post_param(self):
        """
        Load overlay table and post run check parameter located at parameters.

        Returns:
        - DataFrame: The overlaytemplate table and sheet PostRunTestDesc in validation parameter from xlsx.
        """
        param = load_parameters(parmPath=self.parmPath)

        overlay_param = self._load_overlay_parm(param=param)

        validate_param = self._load_valid_param(param=param)

        model_param = self._load_model_param(param=param)

        return overlay_param, validate_param, model_param


class post_run_validation(post_run_validation_basic):
    def __init__(self, context):
        super().__init__(context)

    def load_basic_data_post(self):
        """
        Load ecl results and instrument data table via call func in post_run_validation_basic.

        Returns:
        - tuple: A tuple containing the loaded data table.
        """
        df_ecl = self.load_ecl_result()
        df_instr = self.load_instr_data()
        param_ov, param_valid, param_model = self.load_post_param()
        return df_ecl, df_instr, param_ov, param_valid, param_model

    # 1. ECL must be >= 0
    def ECL_check(self, df_ecl):
        ecl_result = df_ecl.copy().reset_index()
        ecl_cols = [x for x in ecl_result.columns.tolist() if ('ECL_FINAL' in x)
                    and ('ON' not in x) and ('OFF' not in x)]
        ecl_check = np.where(ecl_result[ecl_cols] < 0)

        if len(ecl_check[0]) == 0:
            return ['Pass']
        else:
            neg_rec = {}
            for i in range(len(ecl_check[0])):
                key = ecl_result.loc[ecl_check[0][i], 'CONTRACT_ID'] + " " + ecl_result.loc[ecl_check[0]
                                                                                            [i], 'ON_OFF_BAL_IND'] + " " + ecl_result.loc[ecl_check[0][i], 'INSTR_TYPE']
                val = ecl_cols[ecl_check[1][i]]
                neg_rec[key] = val
            return ['Fail', neg_rec]

    # 2. Lifetime ECL must be > ECL 12m
    def ECL_LT_12M_check(self, df_ecl):
        ecl_result = df_ecl.copy().reset_index()
        engin_res = np.where(
            ecl_result['ECL_ENGINE_12M_OCY'] > ecl_result['ECL_ENGINE_LT_OCY'])

        if len(engin_res[0]) == 0:
            return ['Pass']
        else:
            engin_fail_rec = []
            for i in range(len(engin_res[0])):
                key = ecl_result.loc[engin_res[0][i], 'CONTRACT_ID'] + \
                    " " + ecl_result.loc[engin_res[0][i], 'ON_OFF_BAL_IND']
                engin_fail_rec = engin_fail_rec.append(key)
            return ['Fail', engin_fail_rec]

    # 3-8. Check Contract_id, prin, acru_int, fee input output consistency

    def consistency_six_chk(self, df_ecl, df_instr):
        ecl_result = df_ecl.copy().reset_index()
        instr_raw = df_instr.copy().reset_index()
        FailRec = []
        ind_dict = {'CONTRACT_ID': 'Pass', 'DRAWN_BAL_OCY': 'Pass',
                    'PRIN_BAL_OCY': 'Pass', 'ACRU_INT_OCY': 'Pass',
                    'PENALTY_OCY': 'Pass', 'OTHER_FEE_AND_CHARGES_OCY': 'Pass'}

        # CONTRACT_ID #TODO: identify missing from on or off bal
        contract_id_input_set = set(instr_raw['CONTRACT_ID'])
        contract_id_output_set = set(ecl_result['CONTRACT_ID'])
        if contract_id_input_set != contract_id_output_set:
            ind_dict['CONTRACT_ID'] = 'Fail'
            if len(list(contract_id_input_set - contract_id_output_set)) > 0:
                for i in list(contract_id_input_set - contract_id_output_set):
                    FailRec.append(['contract_id (missing): ' + str(i)])
            if len(list(contract_id_output_set - contract_id_input_set)) > 0:
                for i in list(contract_id_output_set - contract_id_input_set):
                    FailRec.append(['contract_id (extra): ' + str(i)])

        # Draw_bal, Prin, Acru_int, Fee
        input_set = instr_raw[['CONTRACT_ID', 'DRAWN_BAL_OCY', 'PRIN_BAL_OCY', 'ACRU_INT_OCY',
                               'PENALTY_OCY', 'OTHER_FEE_AND_CHARGES_OCY']].copy().drop_duplicates().set_index('CONTRACT_ID')
        output_set = ecl_result[['CONTRACT_ID', 'DRAWN_BAL_OCY', 'PRIN_BAL_OCY', 'ACRU_INT_OCY',
                                 'PENALTY_OCY', 'OTHER_FEE_AND_CHARGES_OCY']].copy().drop_duplicates().set_index('CONTRACT_ID')
        input_set.columns = [x+'_IN' for x in input_set.columns]
        output_set.columns = [x+'_OUT' for x in output_set.columns]
        # 20231106: fill nan with 0 to ensure the comparison
        input_set = input_set.fillna(0)
        output_set = output_set.fillna(0)
        # Use default inner join to exclude non-matching CONTRACT_ID
        comp_df = pd.merge(input_set, output_set,
                           left_index=True, right_index=True)
        for i in ['DRAWN_BAL_OCY', 'PRIN_BAL_OCY', 'ACRU_INT_OCY', 'PENALTY_OCY', 'OTHER_FEE_AND_CHARGES_OCY']:
            comp_df['DIFF_' +
                    i] = round((comp_df[i+'_IN']-comp_df[i+'_OUT']), 2) == 0.0
            if False in list(comp_df['DIFF_' + i].unique()):
                ind_dict[i] = 'Fail'
                for j in comp_df.index[comp_df['DIFF_' + i] == False].to_list():
                    FailRec.append([i+'_mismatch: ' + j])

        return [ind_dict, FailRec]

    # 9. Result range check
    def _range_chk(self, df_ecl, col, cond):
        result = df_ecl.copy().reset_index()
        if cond == '>0':
            res_ind = np.where(result[col] < 0)
        elif cond == '0-1':
            res_ind1 = np.where(result[col] < 0)
            res_ind2 = np.where(result[col] > 1)
            res_ind = res_ind1+res_ind2
        else:
            return ['Wrong Condition']

        na_ind = np.where(result[col].isnull())
        if len(res_ind[0]) == 0 and len(na_ind[0]) == 0:
            return ['Pass']
        else:
            res_list = []
            if len(res_ind[0]) > 0:
                for i in range(len(res_ind[0])):
                    res_list.append(
                        result.loc[res_ind[0][i], 'CONTRACT_ID'] + ': range check failed')
            if len(na_ind[0]) > 0:
                for i in range(len(na_ind[0])):
                    res_list.append(
                        result.loc[na_ind[0][i], 'CONTRACT_ID'] + ': missing')
            return list(set(res_list))

    def range_chk_all(self, df_ecl):
        range_chk_all = {}
        # EFF_INT_RT Value > 0
        range_chk_all['EFF_INT_RT'] = self._range_chk(
            df_ecl, 'EFF_INT_RT', '>0')
        # EAD_OCY: Value > 0
        range_chk_all['EAD_TOT_OCY'] = self._range_chk(df_ecl, 'EAD_OCY', '>0')
        # IFRS9_PD_12M: Value between 0 - 1
        range_chk_all['IFRS9_PD_12M'] = self._range_chk(
            df_ecl, 'IFRS9_PD_12M', '0-1')
        # IFRS9_PD_12M_MADJ: Value between 0 - 1
        range_chk_all['IFRS9_PD_12M_MADJ'] = self._range_chk(
            df_ecl, 'IFRS9_PD_12M_MADJ', '0-1')
        # IFRS9_PD_LT: Value between 0 - 1
        range_chk_all['IFRS9_PD_LT'] = self._range_chk(
            df_ecl, 'IFRS9_PD_LT', '0-1')
        # IFRS9_LGD_12M: Value between 0 - 1
        range_chk_all['IFRS9_LGD_12M'] = self._range_chk(
            df_ecl, 'IFRS9_LGD_12M', '0-1')

        range_chk_rec = {}
        for key, val in range_chk_all.items():
            if val != ['Pass']:
                range_chk_rec[key] = val

        if any(range_chk_rec) is False:
            return ['Pass']
        else:
            return ['Fail', range_chk_rec]

    # 10. No STAGE_FINAL missing
    def stage_final_check(self, df_ecl):
        result = df_ecl.copy().reset_index()
        stage_final_res = np.where(result['STAGE_FINAL'].isnull())
        if len(stage_final_res[0]) == 0:
            return ['Pass']
        else:
            engin_fail_rec = []
            for i in range(len(stage_final_res[0])):
                key = result.loc[stage_final_res[0][i], 'CONTRACT_ID']
                engin_fail_rec.append(key)
            engin_fail_rec = list(set(engin_fail_rec))
            return ['Fail', engin_fail_rec]

    # 11. Overlay
    def ecl_override_check(self, df_ecl, param_ov):

        res_chk_df = df_ecl[['CONTRACT_ID', 'DATA_SOURCE_CD',
                             'EAD_OCY', 'STAGE_FINAL', 'ECL_FINAL_OCY']].copy()
        df_mo = param_ov['Overlay'][['CONTRACT_ID',
                                     'STAGE_OVR', 'ECL_RATE_OVR', 'ECL_OVR']].copy()
        df_mo = df_mo.rename(columns={'STAGE_OVR': 'STAGE_MO',
                                      'ECL_RATE_OVR': 'ECL_RATE_MO',
                                      'ECL_OVR': 'ECL_MO'})
        df_override = df_mo.copy().reset_index()

        # merge by contract_id
        for_check_df = pd.merge(df_override, res_chk_df,
                                how='right', on=['CONTRACT_ID'])
        for_check_df['ECL_RATE_MO'] = for_check_df['ECL_RATE_MO'] * \
            for_check_df['EAD_OCY']
        for_check_df['ECL_MANUAL_FINAL'] = np.where(for_check_df['ECL_MO'].isnull(
        ), for_check_df['ECL_RATE_MO'], for_check_df['ECL_MO'])

        # fill not overlay value with final value
        for_check_df['ECL_MO_OCY'] = for_check_df['ECL_MANUAL_FINAL'].fillna(
            for_check_df['ECL_FINAL_OCY'])
        for_check_df['STAGE_MO'] = for_check_df['STAGE_MO'].fillna(
            for_check_df['STAGE_FINAL'])

        for_check_df['ECL_COMPARE'] = (abs(
            for_check_df.ECL_FINAL_OCY - for_check_df.ECL_MO_OCY) >= 1).replace({True: 1, False: 0})
        for_check_df['STAGE_COMPARE'] = (
            for_check_df.STAGE_FINAL != for_check_df.STAGE_MO).replace({True: 1, False: 0})
        if for_check_df['ECL_COMPARE'].sum() + for_check_df['STAGE_COMPARE'].sum() == 0:
            return ['Pass']
        else:
            engin_fail_rec = []
            res_final = for_check_df.query(
                'ECL_COMPARE == 1 or STAGE_COMPARE == 1').copy().reset_index(drop=True)
            for i in range(res_final.shape[0]):
                if res_final.loc[i, 'ECL_COMPARE'] == 1:
                    key = res_final.loc[i, 'CONTRACT_ID'] + \
                        ': ECL override mismatch'
                    engin_fail_rec.append(key)
                if res_final.loc[i, 'STAGE_COMPARE'] == 1:
                    key = res_final.loc[i, 'CONTRACT_ID'] + \
                        ': Stage override mismatch'
                    engin_fail_rec.append(key)
            return ['Fail', engin_fail_rec]


class post_run_validation_report(post_run_validation):
    def __init__(self, context):
        super().__init__(context)

    def post_run_validation_report_out(self):
        outdf_dict = {}
        # run preparation steps
        df_ecl_all, df_instr, param_ov, validparm, modelparam = self.load_basic_data_post()
        df_ecl = df_ecl_all.query("ECL_APPROACH == 'COLLECTIVE_ASSESSMENT'")
        # run all validation
        try:
            stepinfo = "ECL results check"
            print(stepinfo)
            # instr table contains ead = 0, filter out them to check.
            df_ecl_ca_on = df_ecl.query(
                'ON_OFF_BAL_IND == "ON" and DRAWN_BAL_OCY > 0 and STAGE_FINAL != 3')
            ecl_res = self.ECL_check(df_ecl=df_ecl_ca_on)
        except Exception as e:
            print(f"An error occured when running {stepinfo}: {e}")

        try:
            stepinfo = "ECL Lifetime >= ECL 12M Check"
            print(stepinfo)
            ecl_lt_12m_res = self.ECL_LT_12M_check(df_ecl=df_ecl)
        except Exception as e:
            print(f"An error occured when running {stepinfo}: {e}")

        try:
            stepinfo = "Consistency Check"
            print(stepinfo)
            # TODO: update to all
            df_ecl_on = df_ecl_all.query(
                "ON_OFF_BAL_IND == 'ON' and DATA_SOURCE_CD == 'LOAN'")
            df_instr_on = df_instr.query(
                "ON_OFF_BAL_IND == 'ON' and DATA_SOURCE_CD == 'LOAN'")
            consistency_res = self.consistency_six_chk(
                df_ecl=df_ecl_on, df_instr=df_instr_on)
            # consistency_res = self.consistency_six_chk(df_ecl=df_ecl,df_instr=df_instr)
        except Exception as e:
            print(f"An error occured when running {stepinfo}: {e}")

        try:
            stepinfo = "Result Range Check"
            print(stepinfo)
            df_ecl_ca_on = df_ecl.query(
                "ON_OFF_BAL_IND == 'ON' and DRAWN_BAL_OCY > 0 and STAGE_FINAL != 3")
            range_res = self.range_chk_all(df_ecl_ca_on)
        except Exception as e:
            print(f"An error occured when running {stepinfo}: {e}")

        try:
            stepinfo = "Stage Final Check"
            print(stepinfo)
            stage_res = self.stage_final_check(df_ecl)
        except Exception as e:
            print(f"An error occured when running {stepinfo}: {e}")

        try:
            stepinfo = "ECL overlay check"
            print(stepinfo)
            ecl_ovr_res = self.ecl_override_check(df_ecl, param_ov)
        except Exception as e:
            print(f"An error occured when running {stepinfo}: {e}")

        # concat to different sheets
        # overall result
        validResCols = ['Post-Run Validation Test', 'Result Flag']
        validresall = [['ECL Check', ecl_res[0]],
                       ['ECL Lifetime vs. ECL 12M Check', ecl_lt_12m_res[0]],
                       ['Contract ID Consistency Check',
                           consistency_res[0]['CONTRACT_ID']],
                       ['Drawn Balance Consistency Check',
                           consistency_res[0]['DRAWN_BAL_OCY']],
                       ['Principal Balance Consistency Check',
                           consistency_res[0]['PRIN_BAL_OCY']],
                       ['Accural Interest Consistency Check',
                           consistency_res[0]['ACRU_INT_OCY']],
                       ['Penalty Fee Consistency Check',
                           consistency_res[0]['PENALTY_OCY']],
                       ['Other Charges Consistency Check',
                           consistency_res[0]['OTHER_FEE_AND_CHARGES_OCY']],
                       ['Result Range Check', range_res[0]],
                       ['Stage Final Check', stage_res[0]],
                       ['ECL Overlay Check', ecl_ovr_res[0]]
                       ]
        validReport = pd.DataFrame(data=validresall, columns=validResCols)
        # concat with test description
        validReport = validReport.merge(
            validparm['PostRunTestDesc'].iloc[:, 1:], how='left', on='Post-Run Validation Test')
        outdf_dict['validReport'] = validReport
        
        # failed test & exception
        FailResCols = ['Post-Run Validation Test',
                       'Result Flag', 'Fail Reason']
        if len(ecl_res) > 1:
            exceptionresall = []
            for key in ecl_res[1].keys():
                exceptionresall.append(
                    ['ECL Check', ecl_res[0], key+" : "+ecl_res[1][key]])
            failReport = pd.DataFrame(
                data=exceptionresall, columns=FailResCols)
            outdf_dict['failReport_ECL'] = failReport
        if len(ecl_lt_12m_res) > 1:
            exceptionresall = []
            for i in ecl_lt_12m_res[1]:
                exceptionresall.append(
                    ['ECL Lifetime vs. ECL 12M Check', ecl_lt_12m_res[0], i])
            failReport = pd.DataFrame(
                data=exceptionresall, columns=FailResCols)
            outdf_dict['failReport_ECL LifeTime12M'] = failReport
        if len(consistency_res[1]) > 0:
            exceptionresall = []
            for i in consistency_res[1]:
                exceptionresall.append(['Consistency Check', 'Fail', i])
            failReport = pd.DataFrame(
                data=exceptionresall, columns=FailResCols)
            outdf_dict['failReport_InOutConsistency'] = failReport
        if len(range_res) > 1:
            exceptionresall = []
            for key, val in range_res[1].items():
                for i in val:
                    exceptionresall.append(
                        ['Result Range Check'+'('+key+')', 'Fail', i])
            failReport = pd.DataFrame(
                data=exceptionresall, columns=FailResCols)
            outdf_dict['failReport_Range'] = failReport
        if len(stage_res) > 1:
            exceptionresall = []
            for key in stage_res[1]:
                exceptionresall.append(
                    ['Stage Final Check'+'('+key+')', 'Fail', i])
            failReport = pd.DataFrame(
                data=exceptionresall, columns=FailResCols)
            outdf_dict['failReport_Stage'] = failReport
        if len(ecl_ovr_res) > 1:
            exceptionresall = []
            for key in ecl_ovr_res[1]:
                exceptionresall.append(
                    ['ECL Overlay Check'+'('+key+')', 'Fail', i])
            failReport = pd.DataFrame(
                data=exceptionresall, columns=FailResCols)
            outdf_dict['failReport_Overlay'] = failReport
        print(outdf_dict)
        return outdf_dict

    # Validation result
    @profile
    def run(self):
        return self.post_run_validation_report_out()
