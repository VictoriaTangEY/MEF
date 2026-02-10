import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import itertools
import json


def _deal_sheet(df, _parm):
    """
    Casts the data types of columns in a parameter table DataFrame `df`,
    based on the specified data types provided in the row named 'Parameter data type'.

    Parameters:
    - df (DataFrame): The input DataFrame containing variable parameter table.
    - _parm (str): Parameter table sheet name.

    Returns:
    - DataFrame: Processed DataFrame with columns' data types converted according to the 'Parameter data type' row.
    """
    col_idx = df.iloc[:, 0].values.tolist()
    try:
        start_idx = col_idx.index(1)
    except Exception as e:
        print(f"Issue observed: no variable count number is 1 in {_parm}")
        print(f"Error as {e}")

    variable_params = df.iloc[start_idx:, :]

    try:
        type_idx = col_idx.index('Parameter data type')
    except Exception as e:
        print(
            f"Issue observed in finding data type row indicator 'Parameter data type'in {_parm} ")
        print(f"Error as {e}")
    param_data_types = df.iloc[type_idx]

    for col in variable_params.columns:
        data_type = param_data_types[col]
        try:
            if data_type == 'Parameter data type':
                # variable_params.loc[:, col] = variable_params.loc[:, col].astype(str)
                variable_params[col] = variable_params[col].apply(
                    lambda x: str(x))
            elif data_type == 'integer':
                # variable_params.loc[:, col] = variable_params.loc[:, col].astype(int)
                variable_params[col] = variable_params[col].apply(
                    lambda x: int(x))
            else:
                variable_params[col] = variable_params[col].astype(data_type)
        except Exception as e:
            print(f"Issue observed in data type {data_type} in {_parm}")
            print(f"Error as {e}")
    return variable_params


def load_parameters(parmPath):
    """
    Load parameter data from Excel files in a specified path and process each sheet to convert data types.

    Parameters:
    - parmPath (str): Path to the directory containing Excel files with parameter data.

    Returns:
    - Dict: A dictionary containing processed parameter data from each sheet, where the keys are sheet names.
    """
    param = dict()

    for _wbFile in itertools.chain(parmPath.glob('[!~]*Param.xlsx'), parmPath.glob('[!~]*Template.xlsx')):
        _sheetList = load_workbook(_wbFile).sheetnames
        for _parm in _sheetList:
            df = pd.read_excel(_wbFile, sheet_name=_parm)
            try:
                param[_parm] = _deal_sheet(df, _parm)
            except Exception as e:
                print(f"Issue observed in sheet {_parm}")
                print(f"Error as {e}")
    return param


def load_configuration_file(configPath):
    """
    Load and parse a JSON configuration file for ECL run background settings.

    Parameters:
    - configPath (str): Path to the JSON configuration file.

    Returns:
    - Dict: Dictionary containing the parsed configuration settings for ECL run background.
    """
    with open(Path(configPath), 'r') as fp:
        c = json.load(fp)

    return c


if __name__ == "__main__":
    configPath = Path(
        r'C:\Users\WH947CH\Engagement\Khan Bank\03_ECL_engine\02_Development\khb_engine\run_config_file.json')
    c = load_configuration_file(configPath=configPath)

    print(c)
    # parmPath = Path(r'C:\Users\WH947CH\Engagement\Khan Bank\99_Data_Server\parameter\20240331')
    # parameters = load_parameters(parmPath)
    # print(parameters.keys())
