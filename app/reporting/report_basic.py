import pandas as pd
from openpyxl import load_workbook
from input_handler.load_parameters import load_parameters
from input_handler.data_preprocessor import data_preprocessor
from datetime import datetime


class env_setting():
    def __init__(self, context):
        self.run_yymm = context.run_yymm
        self.prtflo_scope = context.prtflo_scope
        self.scenario_version = context.SCENARIO_VERSION
        self.scenario_set = context.scenario_set
        self.days_in_year = context.days_in_year
        self.days_in_month = context.days_in_month
        self.total_yr = context.total_yr

        self.masterPath = context.masterPath
        self.parmPath = context.parmPath
        self.dataPathScen = context.dataPathScen
        self.inDataPath = context.inDataPath
        self.previnDataPath = context.previnDataPath
        self.resultPath = context.resultPath
        self.prevResultPath = context.prevResultPath
        self.dataName = context.dataNameScen  # scenario data

        self.dtype_tbl = context.dtype_tbl
        self.inputDataExtECL = context.inputDataExtECL
        self.instrument_table_name = context.instrument_table_name
        self.exchange_rate_table_name = context.exchange_rate_table_name
        self.repayment_table_name = context.repayment_table_name
        self.sa_fs_table_name = context.sa_fs_table_name
        self.sa_other_debt_table_name = context.sa_other_debt_table_name
        # self.collateral_table_name = context.collateral_table_name

        self.mute_eir = context.mute_eir
        self.mute_stage_consistency = context.mute_stage_consistency


class report_load_info(env_setting):
    def __init__(self, context):
        super().__init__(context)

    def load_report_template(self, parmPath):
        reporttemp = dict()
        _file = list(self.parmPath.glob(f'reportingForm.xlsx'))[0]
        _sheetList = load_workbook(_file).sheetnames
        for _parm in _sheetList:
            _key = _parm.split(' ')[-1]
            reporttemp[_key] = pd.read_excel(_file, sheet_name=_parm)
        return reporttemp

    def load_all_param(self):
        param = load_parameters(parmPath=self.parmPath)
        return param

    def load_instr_data(self, param):
        dp = data_preprocessor(context=self)

        exposure_std = dp.standardize_data(inDataPath=self.inDataPath,
                                           rawDataName=self.instrument_table_name,
                                           dtype_tbl=self.dtype_tbl,
                                           param_dict=param,
                                           inputDataExt=self.inputDataExtECL)
        return exposure_std[0]

    def load_prev_instr_data(self, param):
        dp = data_preprocessor(context=self)

        prev_exposure_std = dp.standardize_data(inDataPath=self.previnDataPath,
                                           rawDataName=self.instrument_table_name,
                                           dtype_tbl=self.dtype_tbl,
                                           param_dict=param,
                                           inputDataExt=self.inputDataExtECL)
        return prev_exposure_std[0]
    
    def load_Assumptions(self, end):
        dp = data_preprocessor(context=self)
        data = dp.load_scenario_data(
            data_path=self.dataPathScen, file_pattern='MEF')
        end_date = datetime.strptime(end, '%Y/%m/%d')
        end_plus_one_year = end_date.replace(
            year=end_date.year + 1).strftime('%Y/%m/%d')
        assumption = data[data['Code:'] == end_plus_one_year]
        return assumption

    def load_ECL_result(self):
        result_path = self.resultPath/'ECL_calculation_result_files_deal_all.csv'
        pre_result_path = self.prevResultPath/'ECL_calculation_result_files_deal_all.csv'
        result_df = pd.read_csv(result_path)
        pre_result_df = pd.read_csv(pre_result_path)
        return result_df, pre_result_df


class report_cond_basic(env_setting):
    def unit_represent_num(self, value, df_conditions):
        """
        Numerical and decimal representation
        df_conditions: report parameters for each report table
        """
        format_str = df_conditions.query(
            "CONDITION == 'Numerical representation'").SETTING.tolist()[0]
        decimal_int = df_conditions.query(
            "CONDITION == 'Decimal representation'").SETTING.tolist()[0]
        if format_str == 'Percentage':
            return f"{value * 100:.{decimal_int}f}%"
        if format_str == 'Absolute value':
            return f"{value:.{decimal_int}f}%"

    def overall_filter(self, df_conditions, df_input):
        """
        Filter table by condition in reporting parameter. 
        df_conditions: report parameters for each report table
        df_input: input table need do filter

        return dataframe
        """
        filter_str = df_conditions.query(
            "CONDITION == 'Overall filter'").SETTING.tolist()[0]
        if filter_str == 'Not applicable':
            return df_input
        else:
            try:
                df_ = df_input.query(filter_str)
            except Exception as e:
                print(
                    f"An error occured when conducting overall filter on input table: {e}")
            return df_

    def set_scaling_number(self, df_conditions):
        """
        Filter table by condition in reporting parameter. 
        df_conditions: report parameters for each report table
        df_input: input table need do filter

        return dataframe
        """
        str_ = df_conditions.query(
            "CONDITION == 'Numerical scaling'").SETTING.tolist()[0]
        return str_

    def scaling_number_category(self, str_):
        """
        Filter table by condition in reporting parameter. 
        df_conditions: report parameters for each report table
        df_input: input table need do filter

        return dataframe
        """
        if str_ == '1000':
            return "thousands"
        elif str_ == '100':
            return "hundreds"
        elif str_ == '1000000':
            return "millions"
        else:
            return str_

    def scaling_num(self, value, df_conditions):
        """
        Numerical and decimal representation
        df_conditions: report parameters for each report table
        """
        format_str = df_conditions.query(
            "CONDITION == 'Numerical representation'").SETTING.tolist()[0]
        numerical_int = df_conditions.query(
            "CONDITION == 'Numerical scaling'").SETTING.tolist()[0]
        if format_str == 'Absolute value':
            return value / int(numerical_int)

    def set_report_date(self, df_conditions):
        """
        load report date from configuration file , example: 20240630
        df_conditions: report parameters for each report table
        return example: '20240630'
        """
        filter_str = df_conditions.query(
            "CONDITION == 'Reporting date'").SETTING.tolist()[0]
        rp_date = getattr(self, filter_str)
        date_ = str(rp_date)
        return date_

    def format_date_text(self, date_):
        """
        date_: str, standarized as string, example: '20240630'

        return: example: 60 June 2024
        """
        date_obj = datetime.strptime(date_, '%Y%m%d')
        formatted_date = date_obj.strftime('%d %B %Y')
        return (formatted_date)

    def format_date_slash(self, date_):
        """
        date_: str or int, standarized as string, example: '20241231' or 20241231

        return: example: '2024/12/31'
        """
        date_str = str(date_)
        if len(date_str) == 8:
            return f'{date_str[:4]}/{date_str[4:6]}/{date_str[6:8]}'
        return date_str

    def set_currency(self, df_conditions):
        """
        load currency parameter setting table , example: 'MNT'
        df_conditions: report parameters for each report table
        return curr_ , 'MNT'
        """
        curr_str = df_conditions.query(
            "CONDITION == 'Currency'").SETTING.tolist()[0]
        curr_ = curr_str.upper()
        return curr_
