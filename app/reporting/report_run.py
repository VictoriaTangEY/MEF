import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from pathlib import Path
from reporting import report_basic
from reporting import rp_master_scale, rp_forward_looking, rp_asset_info, rp_pd_range_disclosure, rp_collateral_movement, rp_concentration_risk_sector, rp_asset_movement_disclosure, rp_ecl_stage_report, rp_bond_rating_dist, rp_ecl_movement_loan_off,rp_ecl_movement_loan_on, rp_bond_breakdown, rp_commentary
from input_handler.load_parameters import load_parameters
import os


def get_month_list_from_folders(production_path, prev_yymm, run_yymm):
    prev_yymm = str(prev_yymm)
    run_yymm = str(run_yymm)
    all_folders = [name for name in os.listdir(production_path) if os.path.isdir(os.path.join(production_path, name))]
    valid_folders = [f for f in all_folders if f.isdigit() and len(f) == 8]
    filtered = [f for f in valid_folders if prev_yymm <= f <= run_yymm]
    filtered.sort()
    return filtered

def process_monthly_movements(rc, param_all, report_temp_all, ecl_results, month_list, interim_wb, 
                            report_module, run_function, sheet_name, report_type_name):
    """
    Generic function to process monthly movements for different report types
    
    Args:
        rc: context setting
        param_all: all parameters
        report_temp_all: report template
        ecl_results: dictionary of ECL results by month
        month_list: list of months to process
        interim_wb: interim workbook to save individual sheets
        report_module: module containing the run function
        run_function: function to run (e.g., run_ecl_movement_loan_on, run_ecl_movement_loan_off)
        sheet_name: name of the sheet to extract (e.g., 'ECL_movement_loan_on')
        report_type_name: name for logging (e.g., 'movement', 'ECL movement off')
    
    Returns:
        list of movement sheets
    """
    movement_sheets = []
    for i in range(1, len(month_list)):
        prev_month = month_list[i-1]
        curr_month = month_list[i]
        ecl_result_prev = ecl_results[prev_month]
        ecl_result = ecl_results[curr_month]
        
        # copy reportingForm.xlsx as temp_wb
        temp_wb = load_workbook(rc.parmPath / 'reportingForm.xlsx')
        
        # run the specific report function
        temp_wb = getattr(report_module, run_function)(
            rc, param_all, report_temp_all, ecl_result, ecl_result_prev, temp_wb)
        
        # extract the specific sheet
        sheet = temp_wb[sheet_name]
        movement_sheets.append(sheet)
        
        # create new sheet in interim_wb for this month pair
        tab_name = f"{prev_month}_{curr_month}"
        
        # Check if sheet already exists and remove it to avoid duplicates
        if tab_name in interim_wb.sheetnames:
            interim_wb.remove(interim_wb[tab_name])
        
        interim_sheet = interim_wb.create_sheet(title=tab_name)
        for row in sheet.iter_rows(values_only=False):
            interim_sheet.append([cell.value for cell in row])
        print(f"Calculating {report_type_name} for: {prev_month} to {curr_month}")
    
    return movement_sheets

def calculate_interim_movements(rc, param_all, report_temp_all):
    """
    Calculate interim movement results for ECL_movement_loan_on and ECL_movement_loan_off
    
    Args:
        rc: context setting
        param_all: all parameters
        report_temp_all: report template
    
    Returns:
        tuple: (all_movement_sheets, all_movement_off_sheets, ecl_results, month_list)
            - all_movement_sheets: list of movement sheets for ECL_movement_loan_on
            - all_movement_off_sheets: list of movement sheets for ECL_movement_loan_off
            - ecl_results: dictionary of ECL results by month
            - month_list: list of months processed
    """
    # get month list
    RUN_YYMM = rc.run_yymm
    PREV_YYMM = rc.prev_yymm
    # get production root directory
    production_path = rc.resultPath.parents[2]
    month_list = get_month_list_from_folders(str(production_path), PREV_YYMM, RUN_YYMM)

    # load ECL results in month list
    ecl_results = {}
    for month in month_list:
        file_path = os.path.join(production_path, str(month), 'data', '02_result', 'ECL_calculation_result_files_deal_all.csv')
        ecl_results[month] = pd.read_csv(file_path)
        print(f"Loading ECL result for month: {month}")

    # create interim directories for different report types
    base_interim_dir = os.path.join(production_path, str(RUN_YYMM), 'data', '03_report', 'interim')
    
    # Create directories for each report type
    ecl_movement_loan_on_dir = os.path.join(base_interim_dir, 'ecl_movement_loan_on')
    ecl_movement_loan_off_dir = os.path.join(base_interim_dir, 'ecl_movement_loan_off')
    
    os.makedirs(ecl_movement_loan_on_dir, exist_ok=True)
    os.makedirs(ecl_movement_loan_off_dir, exist_ok=True)
    
    # Process ECL_movement_loan_on monthly movements
    ecl_movement_loan_on_filename = f"{str(PREV_YYMM)}_{str(RUN_YYMM)}.xlsx"
    ecl_movement_loan_on_path = os.path.join(ecl_movement_loan_on_dir, ecl_movement_loan_on_filename)
    
    if os.path.exists(ecl_movement_loan_on_path):
        ecl_movement_loan_on_wb = load_workbook(ecl_movement_loan_on_path)
    else:
        ecl_movement_loan_on_wb = Workbook()
        if 'Sheet' in ecl_movement_loan_on_wb.sheetnames:
            std = ecl_movement_loan_on_wb['Sheet']
            ecl_movement_loan_on_wb.remove(std)
    
    all_movement_sheets = process_monthly_movements(
        rc, param_all, report_temp_all, ecl_results, month_list, ecl_movement_loan_on_wb,
        rp_ecl_movement_loan_on, 'run_ecl_movement_loan_on', 'ECL_movement_loan_on', 'movement'
    )
    
    # save ecl_movement_loan_on interim workbook
    ecl_movement_loan_on_wb.save(ecl_movement_loan_on_path)
    print(f"Saving ecl_movement_loan_on interim workbook: {ecl_movement_loan_on_filename}")
    
    # Process ECL_movement_loan_off monthly movements
    ecl_movement_loan_off_filename = f"{str(PREV_YYMM)}_{str(RUN_YYMM)}.xlsx"
    ecl_movement_loan_off_path = os.path.join(ecl_movement_loan_off_dir, ecl_movement_loan_off_filename)
    
    if os.path.exists(ecl_movement_loan_off_path):
        ecl_movement_loan_off_wb = load_workbook(ecl_movement_loan_off_path)
    else:
        ecl_movement_loan_off_wb = Workbook()
        if 'Sheet' in ecl_movement_loan_off_wb.sheetnames:
            std = ecl_movement_loan_off_wb['Sheet']
            ecl_movement_loan_off_wb.remove(std)
    
    all_movement_off_sheets = process_monthly_movements(
        rc, param_all, report_temp_all, ecl_results, month_list, ecl_movement_loan_off_wb,
        rp_ecl_movement_loan_off, 'run_ecl_movement_loan_off', 'ECL_movement_loan_off', 'ECL movement off'
    )
    
    # save ecl_movement_loan_off interim workbook
    ecl_movement_loan_off_wb.save(ecl_movement_loan_off_path)
    print(f"Saving ecl_movement_loan_off interim workbook: {ecl_movement_loan_off_filename}")
    
    return all_movement_sheets, all_movement_off_sheets, ecl_results, month_list

def run(rc):
    PREV_YYMM = rc.prev_yymm
    RUN_YYMM = rc.run_yymm
    rp_b = report_basic.report_load_info(context=rc)
    rp_f = report_basic.report_cond_basic(context=rc)
    wb = load_workbook(rc.parmPath / 'reportingForm.xlsx')

    # load report parameters
    param_all = rp_b.load_all_param()
    # load report template
    report_temp_all = rp_b.load_report_template(rp_b.parmPath)

    # calculate interim movements
    all_movement_sheets, all_movement_off_sheets, ecl_results, month_list = calculate_interim_movements(
        rc, param_all, report_temp_all)
    ecl_result_prev = ecl_results[str(PREV_YYMM)]
    ecl_result_curr = ecl_results[str(RUN_YYMM)]

    # Master scale
    ecl_result = rp_b.load_ECL_result()
    wb = rp_master_scale.run_master_scale(
        rc, param_all, report_temp_all, ecl_result[0], wb)

    # FL
    t_zero = rc.T_ZERO
    date_obj = datetime.strptime(str(t_zero), "%Y%m%d")
    end = date_obj.strftime("%Y/%m/%d")
    assumption = rp_b.load_Assumptions(end)
    wb = rp_forward_looking.run_forward_looking(
        context=rc, param=param_all, report_temp=report_temp_all, assumption=assumption, wb=wb)

    # asset info
    ecl_result = rp_b.load_ECL_result()
    sum_dic = rp_asset_info.sum_value(rc, param_all, ecl_result[0])
    wb = rp_asset_info.run_asset_info(
        rc, report_temp_all, sum_dic, wb)

    # pd range disclosure
    wb = rp_pd_range_disclosure.run_pd_range_disclosure(
        rc, param_all, report_temp_all, ecl_result, wb)

    # concentration risk sector
    wb = rp_concentration_risk_sector.run_concentration_risk_sector(
        rc, param_all, report_temp_all, ecl_result, wb)

    # collateral movement
    wb = rp_collateral_movement.run_collateral_movement(
        rc, param_all, report_temp_all, ecl_result, wb)
    
    # Asset_movement_disclosure
    wb = rp_asset_movement_disclosure.run_asset_movement_disclosure(
        rc, param_all, report_temp_all, ecl_result, wb)
    
    # ECL Stage report
    wb = rp_ecl_stage_report.run_ecl_stage_report(
        rc, param_all, report_temp_all, ecl_result, wb)
    
    # Bond Rating Dist
    wb = rp_bond_rating_dist.run_bond_rating_dist(
        rc, param_all, report_temp_all, ecl_result[0], wb) 

    # Bond breakdown
    wb = rp_bond_breakdown.run_bond_breakdown(
        rc, param_all, report_temp_all, ecl_result_curr, ecl_result_prev, wb)  

    # ECL_movement_loan_on - sum up all adjacent month movements
    wb = rp_ecl_movement_loan_on.run_ecl_movement_loan_on_summed(
        rc, param_all, report_temp_all, all_movement_sheets, wb)
    
    # ECL_movement_loan_off - sum up all adjacent month movements
    wb = rp_ecl_movement_loan_off.run_ecl_movement_loan_off_summed(
        rc, param_all, report_temp_all, all_movement_off_sheets, wb)
    
    # Commentary
    wb = rp_commentary.run_commentary(
        rc, param_all, report_temp_all, ecl_result_curr, ecl_result_prev, wb)

    return wb
